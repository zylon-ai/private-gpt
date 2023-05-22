# https://github.com/hwchase17/langchain/issues/1862

from openpyxl import load_workbook
from typing import Dict, List, Optional
from langchain.docstore.document import Document
from langchain.document_loaders.base import BaseLoader

class XLSXLoader(BaseLoader):
   """Loads an XLSX file into a list of documents.

   Each document represents one row of the XLSX file. Every row is converted into a
   key/value pair and outputted to a new line in the document's page_content.

   The source for each document loaded from xlsx is set to the value of the
   'file_path' argument for all documents by default.
   You can override this by setting the 'source_column' argument to the
   name of a column in the XLSX file.
   The source of each document will then be set to the value of the column
   with the name specified in 'source_column'.

   Output Example:
       .. code-block:: txt

           column1: value1
           column2: value2
           column3: value3
   """

   def __init__(
           self,
           file_path: str,
           source_column: Optional[str] = None,
           sheet_name: Optional[str] = None,
           encoding: Optional[str] = None,
   ):
      self.file_path = file_path
      self.source_column = source_column
      self.sheet_name = sheet_name
      self.encoding = encoding



   def load(self) -> List[Document]:
    docs = []

    wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)
    ws = wb[self.sheet_name] if self.sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        row_values = [cell.value for cell in row]
        row_dict = dict(zip(headers, row_values))

        # Update this line to fix the error
        content = "\n".join(f"{k.strip()}: {v.strip() if isinstance(v, str) else v}" for k, v in row_dict.items() if v is not None)

        if self.source_column is not None:
            source = row_dict[self.source_column]
        else:
            source = self.file_path
        metadata = {"source": source, "row": i}
        doc = Document(page_content=content, metadata=metadata)
        docs.append(doc)

    return docs
