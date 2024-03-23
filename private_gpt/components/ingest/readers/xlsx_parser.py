import json
from pathlib import Path
from typing import List, Optional, Dict
from openpyxl import load_workbook

from llama_index.core.readers.base import BaseReader
from llama_index.core.schema import Document

class XLSXParser(BaseReader):
    """XLSX parser."""

    def __init__(self) -> None:
        """Init parser."""
        super().__init__()

    def load_data(self, input_file: Path, extra_info: Optional[Dict] = {}) -> List[Document]:
        """Parse file."""
        documents = []  # This will contain the list of Document objects
        wb = load_workbook(filename=str(input_file), read_only=True)
        # loop over all sheets
        for sheet in wb:
            sheet_data = []  # This will contain the list of rows for the current sheet
            keys = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue

                # Initialize keys with the first row (header)
                if row_index == 0:
                    keys = [str(cell) if cell is not None else "" for cell in row]
                    continue

                # Ensure each row has the same number of columns as the header
                row_data = [(keys[i], str(cell) if cell is not None else "") for i, cell in enumerate(row) if i < len(keys)]
                sheet_data.append(dict(row_data))

            # Create a Document object with the sheet data and any extra metadata
            document = Document(text=json.dumps(sheet_data, ensure_ascii=False), metadata=extra_info)
            documents.append(document)

        return documents